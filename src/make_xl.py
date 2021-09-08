from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, PatternFill

# 표 스타일 속성값
THIN_BORDER = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

es_field = [
    'page_show',
    'focus.voice_text.button',
    'click.voice_text.button',
    'page_show_stb',
    'focus.voice_text.button_stb',
    'click.voice_text.button_stb',
]

# 슬롯 수 제거
type_header = ['일자',
               '넛지종류',
               '넛지 노출 수',
               '넛지 포커스 수',
               '넛지 클릭 수',
               '노출 대비 포커스',
               '포커스 대비 클릭',
               '클릭 전환율',
               '노출 STB 수',
               '포커스 STB 수',
               '클릭 STB 수',
               '노출 대비 포커스 STB',
               '포커스 대비 클릭 STB',
               '이용율']

# 구분, 슬롯 수 제거
location_header = ['일자',
                   '넛지위치',
                   '넛지 노출 수',
                   '넛지 포커스 수',
                   '넛지 클릭 수',
                   '노출 대비 포커스',
                   '포커스 대비 클릭',
                   '클릭 전환율',
                   '노출 STB 수',
                   '포커스 STB 수',
                   '클릭 STB 수',
                   '노출 대비 포커스 STB',
                   '포커스 대비 클릭 STB',
                   '이용율']


# 헤더 색상, 글씨 굵게 변경
def header_styles(area):
    for row in area:
        for cell in row:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='b7b9bc', end_color='b7b9bc', fill_type='solid')


# 셀 표 모든테두리
def sheet_styles(full_area):
    for row in full_area:
        for cell in row:
            cell.border = THIN_BORDER


# 전환율, 이용률 구하기
def percentage(name):

    if 'page_show' in name:
        for i in es_field:
            if i not in name:
                name[i] = 0
    else:
        name['click'] = 0
        name['use'] = 0
        name['page_show_focus'] = 0
        name['focus_click'] = 0
        name['page_show_focus_stb'] = 0
        name['focus_click_stb'] = 0
        name['click.voice_text.button_stb'] = 0
        name['click.voice_text.button'] = 0
        name['focus.voice_text.button'] = 0
        name['focus.voice_text.button_stb'] = 0

    if 'page_show' in name:
        # 클릭 전환율
        try:
            name['click'] = f"{(name['click.voice_text.button'] / name['page_show']) * 100 : .2f}%"
        except ZeroDivisionError as zero:
            name['click'] = f'{0 : .2f}%'

        # 이용율
        try:
            name['use'] = f"{(name['click.voice_text.button_stb'] / name['page_show_stb']) * 100 : .2f}%"
        except ZeroDivisionError as zero:
            name['use'] = f'{0 : .2f}%'

        # 노출 대비 포커스
        try:
            name['page_show_focus'] = f"{(name['focus.voice_text.button'] / name['page_show']) * 100 : .1f}%"
        except ZeroDivisionError as zero:
            name['page_show_focus'] = f'{0 : .1f}%'

        # 포커스 대비 클릭
        try:
            name['focus_click'] = f"{(name['click.voice_text.button'] / name['focus.voice_text.button']) * 100 : .0f}%"
        except ZeroDivisionError as zero:
            name['focus_click'] = f'{0 : .0f}%'

        # 노출 대비 포커스 STB
        try:
            name['page_show_focus_stb'] = f"{(name['focus.voice_text.button_stb'] / name['page_show_stb']) * 100 : .1f}%"
        except ZeroDivisionError as zero:
            name['page_show_focus_stb'] = f'{0 : .1f}%'

        # 포커스 대비 클릭 STB
        try:
            name['focus_click_stb'] = f"{(name['click.voice_text.button_stb'] / name['focus.voice_text.button_stb']) * 100 : .0f}%"
        except ZeroDivisionError as zero:
            name['focus_click_stb'] = f'{0 : .0f}%'


def insert_data(location_data, type_data):
    # 시트 번호를 알기위한 변수
    num = 1

    wb = Workbook()
    type_ws = wb.create_sheet(title="종류별 자동넛지 데이터")
    location_ws = wb.create_sheet(title="위치별 자동넛지 데이터")
    test_ws = []

    # 종류별 헤더값 입력
    type_ws.append(type_header)
    # 종류별 헤더 스타일 지정
    header_styles(type_ws['A1:N1'])
    # 종류별 데이터 입력
    for t in type_data:
        percentage(t)

        type_ws.append([
            t['day'],
            t['type_value'],
            t['page_show'],
            t['focus.voice_text.button'],
            t['click.voice_text.button'],
            t['page_show_focus'],
            t['focus_click'],
            t['click'],
            t['page_show_stb'],
            t['focus.voice_text.button_stb'],
            t['click.voice_text.button_stb'],
            t['page_show_focus_stb'],
            t['focus_click_stb'],
            t['use']
        ])

        num = num + 1

    # 셀 테두리
    sheet_styles(type_ws['A1:N' + str(num)])

    # 초기화
    num = 1

    # ----------------------------------------------------------------------------------------------------------------

    # 위치별 헤더값 입력
    location_ws.append(location_header)
    # 종류별 헤더 스타일 지정
    header_styles(location_ws['A1:N1'])
    # 종류별 데이터 입력
    for lo in location_data:
        percentage(lo)

        location_ws.append([
            lo['day'],
            lo['location_value'],
            lo['page_show'],
            lo['focus.voice_text.button'],
            lo['click.voice_text.button'],
            lo['page_show_focus'],
            lo['focus_click'],
            lo['click'],
            lo['page_show_stb'],
            lo['focus.voice_text.button_stb'],
            lo['click.voice_text.button_stb'],
            lo['page_show_focus_stb'],
            lo['focus_click_stb'],
            lo['use']
        ])
        num = num + 1

    # 셀 테두리
    sheet_styles(location_ws['A1:N' + str(num)])

    del wb['Sheet']
    wb.save('자동넛지_통계데이터(취합전).xlsx')
